from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from supabase.client import create_client, Client
import os
from dotenv import load_dotenv
from datetime import datetime, date, timezone, timedelta
import uuid
from typing import Optional
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import tempfile
import pandas as pd

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 한국 시간대 (KST = UTC+9)
KST = timezone(timedelta(hours=9))

def get_korean_date():
    """한국 시간 기준으로 오늘 날짜를 반환"""
    return datetime.now(KST).date()

def get_korean_datetime():
    """한국 시간 기준으로 현재 시간을 반환"""
    return datetime.now(KST)

def process_excel_p3(file_path):
    """P3 정렬 처리 (중복 제거 포함)"""
    logger.info(f"P3 처리 시작: {file_path}")
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    last_row = ws.max_row
    last_col = ws.max_column
    
    logger.info(f"Excel 파일 크기: {last_row}행 x {last_col}열")

    p1p2_dict = {}
    p3_dict = {}

    for i in range(3, last_row + 1):
        b = str(ws.cell(i, 2).value)
        type_code = str(ws.cell(i, 15).value)
        if type_code.startswith("P1") or type_code.startswith("P2"):
            p1p2_dict[b] = True
        elif type_code.startswith("P3"):
            p3_dict.setdefault(b, []).append(i)

    logger.info(f"P1/P2 항목 수: {len(p1p2_dict)}, P3 항목 수: {len(p3_dict)}")

    filtered = {k: v for k, v in p3_dict.items() if k not in p1p2_dict}
    logger.info(f"중복 제거 후 P3 항목 수: {len(filtered)}")

    group_array = []
    for key, rows in filtered.items():
        기준문자 = str(ws.cell(rows[0], 15).value)[2:3]
        group_array.append((key, 기준문자, rows))
    group_array.sort(key=lambda x: x[1])
    
    logger.info(f"정렬된 그룹 수: {len(group_array)}")

    new_wb = openpyxl.Workbook()
    result_ws = new_wb.active
    result_ws.title = "정렬결과"

    result_row = 3
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_bold = Font(color="FF0000", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for key, 기준문자, rows in group_array:
        m_high = any(ws.cell(i, 13).value not in [None, 0, ''] for i in rows)
        n_high = any(ws.cell(i, 14).value not in [None, 0, ''] for i in rows)
        start = result_row

        for i in rows:
            for c in range(1, last_col + 1):
                value = ws.cell(i, c).value
                result_ws.cell(result_row, c).value = value
                result_ws.cell(result_row, c).alignment = center_align
            if len(rows) > 1:
                result_ws.cell(result_row, 2).fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            result_row += 1

        if m_high:
            for r in range(start, result_row):
                cell = result_ws.cell(r, 13)
                cell.fill = yellow_fill
                cell.font = red_bold
                cell.alignment = center_align

        if n_high:
            for r in range(start, result_row):
                cell = result_ws.cell(r, 14)
                cell.fill = yellow_fill
                cell.font = red_bold
                cell.alignment = center_align

        def merge_same(col):
            val = result_ws.cell(start, col).value
            if all(result_ws.cell(r, col).value == val for r in range(start, result_row)):
                result_ws.merge_cells(start_row=start, start_column=col, end_row=result_row - 1, end_column=col)
                result_ws.cell(start, col).alignment = center_align

        merge_same(13)
        merge_same(14)

    for r in range(1, 3):
        for c in range(1, last_col + 1):
            value = ws.cell(r, c).value
            result_ws.cell(r, c).value = value
            result_ws.cell(r, c).alignment = center_align

    for col in range(1, last_col + 1):
        max_len = max(len(str(result_ws.cell(r, col).value or "")) for r in range(1, result_ws.max_row + 1))
        result_ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    return new_wb

def process_excel_general(file_path, 기준코드):
    """일반 정렬 처리 (P1, P2 등)"""
    logger.info(f"{기준코드} 처리 시작: {file_path}")
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    last_row = ws.max_row
    last_col = ws.max_column
    
    logger.info(f"Excel 파일 크기: {last_row}행 x {last_col}열")

    group_dict = {}
    for i in range(3, last_row + 1):
        code = str(ws.cell(i, 15).value)
        if code.startswith(기준코드):
            key = str(ws.cell(i, 2).value)
            group_dict.setdefault(key, []).append(i)

    logger.info(f"{기준코드} 항목 수: {len(group_dict)}")

    group_array = []
    for key, rows in group_dict.items():
        기준문자 = str(ws.cell(rows[0], 15).value)[2:3]
        group_array.append((key, 기준문자, rows))
    group_array.sort(key=lambda x: x[1])
    
    logger.info(f"정렬된 그룹 수: {len(group_array)}")

    new_wb = openpyxl.Workbook()
    result_ws = new_wb.active
    result_ws.title = "정렬결과"

    result_row = 3
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_bold = Font(color="FF0000", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    for _, _, rows in group_array:
        start = result_row
        m_high = any(ws.cell(i, 13).value not in [None, 0, ''] for i in rows)
        n_high = any(ws.cell(i, 14).value not in [None, 0, ''] for i in rows)

        for i in rows:
            for c in range(1, last_col + 1):
                result_ws.cell(result_row, c).value = ws.cell(i, c).value
                result_ws.cell(result_row, c).alignment = center_align
            if len(rows) > 1:
                result_ws.cell(result_row, 2).fill = pink_fill
            result_row += 1

        if m_high:
            for r in range(start, result_row):
                cell = result_ws.cell(r, 13)
                cell.fill = yellow_fill
                cell.font = red_bold
                cell.alignment = center_align

        if n_high:
            for r in range(start, result_row):
                cell = result_ws.cell(r, 14)
                cell.fill = yellow_fill
                cell.font = red_bold
                cell.alignment = center_align

        def merge_same(col):
            val = result_ws.cell(start, col).value
            if all(result_ws.cell(r, col).value == val for r in range(start, result_row)):
                result_ws.merge_cells(start_row=start, start_column=col, end_row=result_row - 1, end_column=col)
                result_ws.cell(start, col).alignment = center_align

        merge_same(13)
        merge_same(14)

    for r in range(1, 3):
        for c in range(1, last_col + 1):
            result_ws.cell(r, c).value = ws.cell(r, c).value
            result_ws.cell(r, c).alignment = center_align

    for col in range(1, last_col + 1):
        max_len = max(len(str(result_ws.cell(r, col).value or "")) for r in range(1, result_ws.max_row + 1))
        result_ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    return new_wb

# 환경 변수 로드 (로컬 개발용)
if os.path.exists('.env'):
    load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'dev-secret-key-change-in-production')

# Supabase 클라이언트 설정
supabase: Optional[Client] = None
try:
    supabase_url = os.getenv('SUPABASE_URL')
    supabase_key = os.getenv('SUPABASE_KEY')
    
    if not supabase_url or not supabase_key:
        logger.error("⚠️ Supabase 환경 변수가 설정되지 않았습니다.")
        raise ValueError("SUPABASE_URL과 SUPABASE_KEY가 필요합니다.")
    
    supabase = create_client(supabase_url, supabase_key)
    logger.info("✅ Supabase 연결 성공")
except Exception as e:
    logger.error(f"❌ Supabase 연결 실패: {e}")
    supabase = None

@app.route('/health')
def health_check():
    return jsonify({'status': 'healthy', 'message': 'WorkTracker is running'})

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Page not found'}), 404

@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # 관리자 로그인 처리
        if 'admin_login' in request.form:
            admin_password = request.form['admin_password']
            
            if admin_password == '0000':
                session['user'] = {
                    'id': 'admin-user-id',
                    'username': 'admin',
                    'role': 'admin'
                }
                logger.info("관리자 로그인 성공")
                return redirect(url_for('admin_dashboard'))
            else:
                return render_template('login.html', error="관리자 비밀번호가 올바르지 않습니다.")
        
        # 일반 사용자 로그인 처리
        username = request.form['username']
        password = request.form['password']
        
        # Supabase 연결 확인
        if not supabase:
            return render_template('login.html', error="데이터베이스 연결에 문제가 있습니다. 관리자에게 문의하세요.")
        
        try:
            # admin 계정 특별 처리
            if username == 'admin' and password == '0000':
                session['user'] = {
                    'id': 'admin-user-id',
                    'username': 'admin',
                    'role': 'admin'
                }
                logger.info("관리자 로그인 성공")
                return redirect(url_for('admin_dashboard'))
            
            # 일반 사용자 로그인 (사용자명만 확인)
            user_data = supabase.table('users').select('*').eq('username', username).execute()
            
            if user_data.data:
                # 실제 구현에서는 비밀번호 해시 검증이 필요합니다
                # 여기서는 간단히 사용자명만 확인
                session['user'] = {
                    'id': user_data.data[0]['id'],
                    'username': user_data.data[0]['username']
                }
                logger.info(f"사용자 로그인 성공: {username}")
                return redirect(url_for('dashboard'))
            else:
                return render_template('login.html', error="사용자명 또는 비밀번호가 올바르지 않습니다.")
        except Exception as e:
            logger.error(f"로그인 에러: {e}")
            return render_template('login.html', error="로그인에 실패했습니다. 다시 시도해주세요.")
    
    # 세션에서 성공 메시지 가져오기
    success_message = session.pop('success_message', None)
    return render_template('login.html', success=success_message)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        department_id = request.form.get('department')
        
        if not supabase:
            return render_template('register.html', error="데이터베이스 연결에 문제가 있습니다. 관리자에게 문의하세요.")
        
        try:
            # 사용자명 중복 검사
            existing_user = supabase.table('users').select('username').eq('username', username).execute()
            
            if existing_user.data:
                return render_template('register.html', error="이미 사용 중인 사용자명입니다. 다른 사용자명을 선택해주세요.")
            
            # 사용자 ID 생성
            user_id = str(uuid.uuid4())
            
            # 사용자 정보 저장
            user_data = {
                'id': user_id,
                'username': username,
                'password_hash': password,  # 실제로는 해시화해야 함
                'created_at': get_korean_datetime().isoformat()
            }
            
            # 소속이 선택된 경우 추가
            if department_id:
                user_data['department_id'] = department_id
            
            result = supabase.table('users').insert(user_data).execute()
            
            if result.data:
                # 성공 메시지를 세션에 저장하고 로그인 페이지로 리다이렉트
                session['success_message'] = '회원가입이 완료되었습니다! 로그인해주세요.'
                return redirect(url_for('login'))
            else:
                return render_template('register.html', error="회원가입에 실패했습니다. 다시 시도해주세요.")
                
        except Exception as e:
            logger.error(f"회원가입 에러: {e}")
            # 중복 키 오류인 경우 특별 처리
            if "duplicate key value violates unique constraint" in str(e):
                return render_template('register.html', error="이미 사용 중인 사용자명입니다. 다른 사용자명을 선택해주세요.")
            return render_template('register.html', error="회원가입 중 오류가 발생했습니다. 다시 시도해주세요.")
    
    # GET 요청 시 소속 목록 조회
    try:
        if supabase:
            departments = supabase.table('departments').select('*').execute().data
        else:
            departments = []
    except Exception as e:
        logger.error(f"소속 목록 조회 에러: {e}")
        departments = []
    
    return render_template('register.html', departments=departments)

@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user']['id']
    username = session['user']['username']
    
    # Supabase 연결 확인
    if not supabase:
        return render_template('error.html', error="데이터베이스 연결에 문제가 있습니다.")
    
    try:
        # 관리자 계정 특별 처리
        if username == 'admin':
            return render_template('admin_dashboard.html', user={'username': 'admin', 'role': 'admin'})
        
        # 일반 사용자 정보와 소속 정보를 함께 조회
        user_info = supabase.table('users').select('*, departments(name)').eq('id', user_id).execute()
        
        if user_info.data:
            user = user_info.data[0]
            user_data = {
                'id': user['id'],
                'username': user['username'],
                'role': user['role'],
                'department_id': user['department_id'],
                'department_name': user['departments']['name'] if user['departments'] else None,
                'created_at': user['created_at']
            }
            
            if user['role'] == 'admin':
                return render_template('admin_dashboard.html', user=user_data)
            else:
                return render_template('dashboard.html', user=user_data)
        else:
            return redirect(url_for('login'))
    except Exception as e:
        logger.error(f"대시보드 에러: {e}")
        return redirect(url_for('login'))

@app.route('/admin_dashboard')
def admin_dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user']['id']
    username = session['user']['username']
    
    # Supabase 연결 확인
    if not supabase:
        return render_template('error.html', error="데이터베이스 연결에 문제가 있습니다.")
    
    # 관리자 권한 확인
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', user_id).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return redirect(url_for('dashboard'))
        except:
            return redirect(url_for('dashboard'))
    
    try:
        # 관리자용 데이터 로드
        departments = supabase.table('departments').select('*').execute().data
        users = supabase.table('users').select('*, departments(name)').execute().data
        work_logs = supabase.table('work_logs').select('*, users(username, departments(name))').execute().data
        
        return render_template('admin_dashboard.html', 
                             user={'username': username, 'role': 'admin'},
                             departments=departments,
                             users=users,
                             tasks=work_logs)
    except Exception as e:
        logger.error(f"관리자 대시보드 에러: {e}")
        return render_template('admin_dashboard.html', 
                             user={'username': username, 'role': 'admin'},
                             departments=[],
                             users=[],
                             tasks=[])

@app.route('/test')
def test():
    return jsonify({'message': 'Test endpoint working!', 'status': 'success'})

@app.route('/test_time')
def test_time():
    """한국시간대 테스트 페이지"""
    return render_template('test_time.html')

@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('login'))

# API 엔드포인트들
@app.route('/api/tasks', methods=['GET'])
def get_tasks():
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    user_id = session['user']['id']
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        # 필터 파라미터 처리
        date_filter = request.args.get('date')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 기본 쿼리 (사용자별)
        query = supabase.table('work_logs').select('*').eq('user_id', user_id)
        
        # 날짜 필터 적용
        if date_filter:
            query = query.eq('work_date', date_filter)
        elif start_date and end_date:
            query = query.gte('work_date', start_date).lte('work_date', end_date)
        else:
            # 기본값: 한국 시간 기준 오늘 날짜
            today = get_korean_date().isoformat()
            query = query.eq('work_date', today)
        
        # 정렬 (최신순)
        query = query.order('created_at', desc=True)
        
        result = query.execute()
        return jsonify(result.data)
        
    except Exception as e:
        logger.error(f"업무 목록 조회 에러: {e}")
        return jsonify({'error': '업무 목록 조회에 실패했습니다.'}), 500

@app.route('/api/tasks', methods=['POST'])
def create_task():
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    user_id = session['user']['id']
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        data = request.get_json()
        
        # 필수 필드 검증
        required_fields = ['work_date', 'start_time', 'task_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} 필드가 필요합니다.'}), 400
        
        # 업무 로그 데이터 생성
        task_data = {
            'user_id': user_id,
            'work_date': data['work_date'],
            'start_time': data['start_time'],
            'task_type': data['task_type'],
            'description': data.get('description', ''),
            'status': data.get('status', '진행중'),
            'created_at': get_korean_datetime().isoformat()
        }
        
        result = supabase.table('work_logs').insert(task_data).execute()
        
        if result.data:
            return jsonify(result.data[0]), 201
        else:
            return jsonify({'error': '업무 등록에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"업무 등록 에러: {e}")
        return jsonify({'error': '업무 등록에 실패했습니다.'}), 500

@app.route('/api/tasks/<task_id>', methods=['GET'])
def get_task(task_id):
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    user_id = session['user']['id']
    username = session['user']['username']
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        # 관리자는 모든 업무 조회 가능, 일반 사용자는 자신의 업무만
        if username == 'admin':
            task_result = supabase.table('work_logs').select('*').eq('id', task_id).execute()
        else:
            task_result = supabase.table('work_logs').select('*').eq('id', task_id).eq('user_id', user_id).execute()
        
        if not task_result.data:
            return jsonify({'error': '업무를 찾을 수 없거나 권한이 없습니다.'}), 404
        
        return jsonify(task_result.data[0])
        
    except Exception as e:
        logger.error(f"업무 조회 에러: {e}")
        return jsonify({'error': '업무 조회에 실패했습니다.'}), 500

@app.route('/api/tasks/<task_id>', methods=['PUT'])
def update_task(task_id):
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    user_id = session['user']['id']
    username = session['user']['username']
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        data = request.get_json()
        
        # 필수 필드 검증
        required_fields = ['work_date', 'start_time', 'task_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} 필드가 필요합니다.'}), 400
        
        # 관리자는 모든 업무 수정 가능, 일반 사용자는 자신의 업무만
        if username == 'admin':
            task_result = supabase.table('work_logs').select('*').eq('id', task_id).execute()
        else:
            task_result = supabase.table('work_logs').select('*').eq('id', task_id).eq('user_id', user_id).execute()
        
        if not task_result.data:
            return jsonify({'error': '업무를 찾을 수 없거나 권한이 없습니다.'}), 404
        
        # 업무 수정
        update_data = {
            'work_date': data['work_date'],
            'start_time': data['start_time'],
            'task_type': data['task_type'],
            'description': data.get('description', ''),
            'updated_at': get_korean_datetime().isoformat()
        }
        
        # 종료 시간이 있으면 추가
        if data.get('end_time'):
            update_data['end_time'] = data['end_time']
            update_data['status'] = '완료'
        
        # 완료 내용이 있으면 추가
        if data.get('complete_description'):
            update_data['complete_description'] = data['complete_description']
        
        result = supabase.table('work_logs').update(update_data).eq('id', task_id).execute()
        
        if result.data:
            return jsonify(result.data[0])
        else:
            return jsonify({'error': '업무 수정에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"업무 수정 에러: {e}")
        return jsonify({'error': '업무 수정에 실패했습니다.'}), 500

@app.route('/api/tasks/<task_id>', methods=['DELETE'])
def delete_task(task_id):
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    user_id = session['user']['id']
    username = session['user']['username']
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        # 관리자는 모든 업무 삭제 가능, 일반 사용자는 자신의 업무만
        if username == 'admin':
            task_result = supabase.table('work_logs').select('*').eq('id', task_id).execute()
        else:
            task_result = supabase.table('work_logs').select('*').eq('id', task_id).eq('user_id', user_id).execute()
        
        if not task_result.data:
            return jsonify({'error': '업무를 찾을 수 없거나 권한이 없습니다.'}), 404
        
        # 업무 삭제
        result = supabase.table('work_logs').delete().eq('id', task_id).execute()
        
        if result.data:
            return jsonify({'message': '업무가 삭제되었습니다.'})
        else:
            return jsonify({'error': '업무 삭제에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"업무 삭제 에러: {e}")
        return jsonify({'error': '업무 삭제에 실패했습니다.'}), 500

@app.route('/api/tasks/<task_id>/complete', methods=['PUT'])
def complete_task(task_id):
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    user_id = session['user']['id']
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        data = request.get_json()
        
        # 필수 필드 검증
        if not data.get('end_time'):
            return jsonify({'error': '종료 시간이 필요합니다.'}), 400
        
        # 업무 존재 및 권한 확인
        task_result = supabase.table('work_logs').select('*').eq('id', task_id).eq('user_id', user_id).execute()
        
        if not task_result.data:
            return jsonify({'error': '업무를 찾을 수 없거나 권한이 없습니다.'}), 404
        
        # 업무 완료 처리
        update_data = {
            'end_time': data['end_time'],
            'complete_description': data.get('complete_description', ''),
            'status': '완료',
            'updated_at': get_korean_datetime().isoformat()
        }
        
        result = supabase.table('work_logs').update(update_data).eq('id', task_id).execute()
        
        if result.data:
            return jsonify(result.data[0])
        else:
            return jsonify({'error': '업무 완료 처리에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"업무 완료 처리 에러: {e}")
        return jsonify({'error': '업무 완료 처리에 실패했습니다.'}), 500

# 관리자용 API 엔드포인트들
@app.route('/api/users', methods=['GET'])
def get_users():
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        # 사용자 목록 조회 (소속 정보 포함)
        users = supabase.table('users').select('*, departments(name)').execute().data
        
        # 응답 데이터 형식 변환
        user_list = []
        for user in users:
            user_data = {
                'id': user['id'],
                'username': user['username'],
                'role': user.get('role', 'user'),
                'department_id': user.get('department_id'),
                'department_name': user['departments']['name'] if user['departments'] else None,
                'created_at': user['created_at']
            }
            user_list.append(user_data)
        
        return jsonify(user_list)
        
    except Exception as e:
        logger.error(f"사용자 목록 조회 에러: {e}")
        return jsonify({'error': '사용자 목록 조회에 실패했습니다.'}), 500

@app.route('/api/departments', methods=['GET'])
def get_departments():
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        # 소속 목록 조회
        departments = supabase.table('departments').select('*').order('created_at', desc=True).execute().data
        return jsonify(departments)
        
    except Exception as e:
        logger.error(f"소속 목록 조회 에러: {e}")
        return jsonify({'error': '소속 목록 조회에 실패했습니다.'}), 500

@app.route('/api/departments', methods=['POST'])
def create_department():
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        data = request.get_json()
        name = data.get('name')
        
        if not name:
            return jsonify({'error': '소속명이 필요합니다.'}), 400
        
        # 소속 생성
        department_data = {
            'name': name,
            'created_at': get_korean_datetime().isoformat()
        }
        
        result = supabase.table('departments').insert(department_data).execute()
        
        if result.data:
            return jsonify(result.data[0]), 201
        else:
            return jsonify({'error': '소속 생성에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"소속 생성 에러: {e}")
        return jsonify({'error': '소속 생성에 실패했습니다.'}), 500

@app.route('/api/departments/<dept_id>', methods=['PUT'])
def update_department(dept_id):
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        data = request.get_json()
        name = data.get('name')
        
        if not name:
            return jsonify({'error': '소속명이 필요합니다.'}), 400
        
        # 소속 수정
        result = supabase.table('departments').update({'name': name}).eq('id', dept_id).execute()
        
        if result.data:
            return jsonify(result.data[0])
        else:
            return jsonify({'error': '소속 수정에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"소속 수정 에러: {e}")
        return jsonify({'error': '소속 수정에 실패했습니다.'}), 500

@app.route('/api/departments/<dept_id>', methods=['DELETE'])
def delete_department(dept_id):
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        # 소속 삭제
        result = supabase.table('departments').delete().eq('id', dept_id).execute()
        
        if result.data:
            return jsonify({'message': '소속이 삭제되었습니다.'})
        else:
            return jsonify({'error': '소속 삭제에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"소속 삭제 에러: {e}")
        return jsonify({'error': '소속 삭제에 실패했습니다.'}), 500

# 관리자 페이지 라우트
@app.route('/admin_users')
def admin_users():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return redirect(url_for('dashboard'))
        except:
            return redirect(url_for('dashboard'))
    
    return render_template('admin_users.html', user={'username': username, 'role': 'admin'})

@app.route('/admin_departments')
def admin_departments():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return redirect(url_for('dashboard'))
        except:
            return redirect(url_for('dashboard'))
    
    return render_template('admin_departments.html', user={'username': username, 'role': 'admin'})

@app.route('/missing_response')
def missing_response():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return redirect(url_for('dashboard'))
        except:
            return redirect(url_for('dashboard'))
    
    return render_template('missing_response.html', user={'username': username, 'role': 'admin'})

@app.route('/install')
def install_app():
    """데이터베이스 초기화 페이지"""
    return render_template('install.html')

@app.route('/pwa-install')
def pwa_install():
    """PWA 설치 안내 페이지"""
    return render_template('pwa_install.html')

@app.route('/api/init-database', methods=['POST'])
def init_database():
    """데이터베이스 완전 초기화 (KST 시간대 적용)"""
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        logger.info("🔄 데이터베이스 초기화 시작...")
        
        # 1단계: 기존 데이터 삭제
        try:
            # work_logs 삭제
            supabase.table('work_logs').delete().neq('id', 0).execute()
            logger.info("✅ work_logs 데이터 삭제 완료")
            
            # users 삭제 (admin 제외)
            supabase.table('users').delete().neq('username', 'admin').execute()
            logger.info("✅ users 데이터 삭제 완료")
            
            # departments 삭제
            supabase.table('departments').delete().neq('id', 0).execute()
            logger.info("✅ departments 데이터 삭제 완료")
            
        except Exception as e:
            logger.warning(f"기존 데이터 삭제 중 오류 (무시): {e}")
        
        # 2단계: 기본 데이터 삽입
        try:
            # departments 삽입
            departments_data = [
                {'name': 'B동보충'},
                {'name': 'A지상보충'},
                {'name': 'A지하보충'}
            ]
            
            for dept in departments_data:
                supabase.table('departments').insert(dept).execute()
            logger.info("✅ departments 데이터 삽입 완료")
            
            # admin 계정 확인 및 생성
            admin_check = supabase.table('users').select('*').eq('username', 'admin').execute()
            if not admin_check.data:
                admin_data = {
                    'username': 'admin',
                    'password_hash': '0000',
                    'role': 'admin'
                }
                supabase.table('users').insert(admin_data).execute()
                logger.info("✅ admin 계정 생성 완료")
            else:
                logger.info("✅ admin 계정 이미 존재")
            
            logger.info("✅ 데이터베이스 초기화 완료")
            return jsonify({
                'message': '데이터베이스가 성공적으로 초기화되었습니다.',
                'kst_time': get_korean_datetime().isoformat(),
                'details': {
                    'departments_created': len(departments_data),
                    'admin_account': 'admin/0000'
                }
            })
            
        except Exception as e:
            logger.error(f"기본 데이터 삽입 실패: {e}")
            return jsonify({'error': '기본 데이터 삽입에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"데이터베이스 초기화 실패: {e}")
        return jsonify({'error': '데이터베이스 초기화에 실패했습니다.'}), 500

# 추가 관리자 API 엔드포인트
@app.route('/api/users/<user_id>/department', methods=['PUT'])
def update_user_department(user_id):
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        data = request.get_json()
        department_id = data.get('department_id')
        
        if department_id is None:
            return jsonify({'error': '소속 ID가 필요합니다.'}), 400
        
        # 사용자 소속 업데이트
        result = supabase.table('users').update({'department_id': department_id}).eq('id', user_id).execute()
        
        if result.data:
            return jsonify(result.data[0])
        else:
            return jsonify({'error': '사용자 소속 변경에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"사용자 소속 변경 에러: {e}")
        return jsonify({'error': '사용자 소속 변경에 실패했습니다.'}), 500

@app.route('/api/users/<user_id>/reset-password', methods=['PUT'])
def reset_user_password(user_id):
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        # 비밀번호를 1234로 초기화
        result = supabase.table('users').update({'password_hash': '1234'}).eq('id', user_id).execute()
        
        if result.data:
            return jsonify({'message': '비밀번호가 초기화되었습니다.'})
        else:
            return jsonify({'error': '비밀번호 초기화에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"비밀번호 초기화 에러: {e}")
        return jsonify({'error': '비밀번호 초기화에 실패했습니다.'}), 500

@app.route('/api/users/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        # 사용자 삭제
        result = supabase.table('users').delete().eq('id', user_id).execute()
        
        if result.data:
            return jsonify({'message': '사용자가 삭제되었습니다.'})
        else:
            return jsonify({'error': '사용자 삭제에 실패했습니다.'}), 500
            
    except Exception as e:
        logger.error(f"사용자 삭제 에러: {e}")
        return jsonify({'error': '사용자 삭제에 실패했습니다.'}), 500

# 관리자용 업무 목록 API
@app.route('/api/admin/tasks', methods=['GET'])
def get_admin_tasks():
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        # 오늘 날짜 (한국 시간 기준)
        today = get_korean_date().isoformat()
        logger.info(f"관리자 업무 목록 조회 - 오늘 날짜: {today}")
        
        # 필터 파라미터 처리
        department = request.args.get('department')
        task_type = request.args.get('task_type')
        status = request.args.get('status')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        time_range = request.args.get('time_range')
        
        # 기본 쿼리 (오늘 업무만)
        query = supabase.table('work_logs').select('*, users(username, departments(name))').eq('work_date', today)
        
        # 필터 적용
        if department:
            query = query.eq('users.departments.name', department)
        if task_type:
            query = query.eq('task_type', task_type)
        if status:
            query = query.eq('status', status)
        
        # 기간 필터가 있는 경우 오늘 기준 제한 해제
        if start_date and end_date:
            query = supabase.table('work_logs').select('*, users(username, departments(name))').gte('work_date', start_date).lte('work_date', end_date)
            if department:
                query = query.eq('users.departments.name', department)
            if task_type:
                query = query.eq('task_type', task_type)
            if status:
                query = query.eq('status', status)
        elif start_date:
            query = supabase.table('work_logs').select('*, users(username, departments(name))').gte('work_date', start_date)
            if department:
                query = query.eq('users.departments.name', department)
            if task_type:
                query = query.eq('task_type', task_type)
            if status:
                query = query.eq('status', status)
        elif end_date:
            query = supabase.table('work_logs').select('*, users(username, departments(name))').lte('work_date', end_date)
            if department:
                query = query.eq('users.departments.name', department)
            if task_type:
                query = query.eq('task_type', task_type)
            if status:
                query = query.eq('status', status)
        
        # 정렬 (최신순)
        query = query.order('created_at', desc=True)
        
        result = query.execute()
        return jsonify(result.data)
        
    except Exception as e:
        logger.error(f"관리자 업무 목록 조회 에러: {e}")
        return jsonify({'error': '업무 목록 조회에 실패했습니다.'}), 500

# 관리자 대시보드 통계 API
@app.route('/api/admin/statistics', methods=['GET'])
def get_admin_statistics():
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    if not supabase:
        return jsonify({'error': '데이터베이스 연결에 문제가 있습니다.'}), 500
    
    try:
        today = get_korean_date().isoformat()
        logger.info(f"오늘 날짜 (KST): {today}")
        
        # 총 작업자 수 (회원가입된 모든 사용자)
        total_users = supabase.table('users').select('id').execute()
        total_workers = len(total_users.data)
        logger.info(f"총 작업자 수: {total_workers}")
        
        # 오늘 작업 참여자 수 (오늘 업무를 등록한 사용자)
        today_workers = supabase.table('work_logs').select('user_id').eq('work_date', today).execute()
        today_participants = len(set(task['user_id'] for task in today_workers.data))
        logger.info(f"오늘 작업 참여자 수: {today_participants}")
        
        # 오늘 업무 통계 (오늘 날짜 기준)
        today_tasks = supabase.table('work_logs').select('status').eq('work_date', today).execute()
        logger.info(f"오늘 업무 데이터: {today_tasks.data}")
        
        today_completed_tasks = len([task for task in today_tasks.data if task['status'] == '완료'])
        today_ongoing_tasks = len([task for task in today_tasks.data if task['status'] == '진행중'])
        today_total_tasks = len(today_tasks.data)
        
        logger.info(f"오늘 완료된 업무: {today_completed_tasks}")
        logger.info(f"오늘 진행중인 업무: {today_ongoing_tasks}")
        logger.info(f"오늘 총 업무: {today_total_tasks}")
        
        statistics = {
            'total_workers': total_workers,
            'today_participants': today_participants,
            'completed_tasks': today_completed_tasks,
            'ongoing_tasks': today_ongoing_tasks,
            'total_tasks': today_total_tasks
        }
        
        logger.info(f"반환할 통계: {statistics}")
        return jsonify(statistics)
        
    except Exception as e:
        logger.error(f"관리자 통계 조회 에러: {e}")
        return jsonify({'error': '통계 조회에 실패했습니다.'}), 500

# 미출대응 엑셀 파일 처리 API
@app.route('/api/meechul-process', methods=['POST'])
def meechul_process():
    """미출대응 Excel 파일 처리"""
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    if session['user'].get('role') != 'admin':
        return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    try:
        # 파일 업로드 확인
        if 'file' not in request.files:
            return jsonify({'error': '파일이 업로드되지 않았습니다.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '파일이 선택되지 않았습니다.'}), 400
        
        # 파일 확장자 확인
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Excel 파일만 업로드 가능합니다.'}), 400
        
        # 기준 코드 확인
        기준코드 = request.form.get('criteria', 'P3')
        if 기준코드 not in ['P1', 'P2', 'P3']:
            return jsonify({'error': '올바르지 않은 기준 코드입니다.'}), 400
        
        logger.info(f"미출대응 처리 시작: 파일={file.filename}, 기준코드={기준코드}")
        
        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            file_path = tmp_file.name
        
        logger.info(f"임시 파일 저장 완료: {file_path}")
        
        try:
            # Excel 처리
            logger.info(f"Excel 처리 시작: {기준코드}")
            if 기준코드 == 'P3':
                processed_wb = process_excel_p3(file_path)
            else:
                processed_wb = process_excel_general(file_path, 기준코드)
            
            logger.info("Excel 처리 완료")
            
            # 결과 파일 저장
            output_filename = f"정렬결과_{기준코드}_{get_korean_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(tempfile.gettempdir(), output_filename)
            processed_wb.save(output_path)
            
            logger.info(f"결과 파일 저장 완료: {output_path}")
            
            # 임시 입력 파일 삭제
            os.unlink(file_path)
            logger.info("임시 입력 파일 삭제 완료")
            
            return jsonify({
                'success': True,
                'message': f'{기준코드} 정렬이 완료되었습니다. 파일명: {output_filename}',
                'filename': output_filename,
                'download_url': f'/api/meechul-download/{output_filename}'
            })
            
        except Exception as e:
            # 임시 파일 정리
            if os.path.exists(file_path):
                os.unlink(file_path)
                logger.info("오류 발생으로 임시 파일 정리 완료")
            logger.error(f"Excel 처리 중 오류: {e}")
            raise e
            
    except Exception as e:
        logger.error(f"미출대응 처리 에러: {e}")
        return jsonify({'error': f'파일 처리 중 오류가 발생했습니다: {str(e)}'}), 500

@app.route('/api/meechul-test')
def meechul_test():
    """미출대응 기능 테스트용 엔드포인트"""
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    if session['user'].get('role') != 'admin':
        return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    try:
        # 테스트용 간단한 Excel 파일 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "테스트"
        
        # 헤더 추가
        headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=2, column=col, value=header)
        
        # 테스트 데이터 추가 (P1, P2, P3 코드 포함)
        test_data = [
            ['1', 'ITEM001', 'P1A', '', '', '', '', '', '', '', '', '', '0', '0', 'P1A'],
            ['2', 'ITEM002', 'P2B', '', '', '', '', '', '', '', '', '', '0', '0', 'P2B'],
            ['3', 'ITEM003', 'P3C', '', '', '', '', '', '', '', '', '', '1', '0', 'P3C'],
            ['4', 'ITEM004', 'P3D', '', '', '', '', '', '', '', '', '', '0', '1', 'P3D'],
        ]
        
        for row_idx, row_data in enumerate(test_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 임시 파일로 저장
        test_filename = f"테스트파일_{get_korean_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
        test_path = os.path.join(tempfile.gettempdir(), test_filename)
        wb.save(test_path)
        
        return jsonify({
            'success': True,
            'message': '테스트 파일이 생성되었습니다.',
            'test_file': test_filename,
            'test_path': test_path
        })
        
    except Exception as e:
        logger.error(f"테스트 파일 생성 오류: {e}")
        return jsonify({'error': f'테스트 파일 생성 중 오류가 발생했습니다: {str(e)}'}), 500

@app.route('/api/meechul-download/<filename>')
def meechul_download(filename):
    """미출대응 처리 결과 파일 다운로드"""
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    if session['user'].get('role') != 'admin':
        return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    try:
        file_path = os.path.join(tempfile.gettempdir(), filename)
        if not os.path.exists(file_path):
            return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"파일 다운로드 에러: {e}")
        return jsonify({'error': '파일 다운로드 중 오류가 발생했습니다.'}), 500

@app.route('/api/process-missing-response', methods=['POST'])
def process_missing_response():
    if 'user' not in session:
        return jsonify({'error': '로그인이 필요합니다.'}), 401
    
    # 관리자 권한 확인
    username = session['user']['username']
    if username != 'admin':
        try:
            user_info = supabase.table('users').select('*').eq('id', session['user']['id']).execute()
            if not user_info.data or user_info.data[0]['role'] != 'admin':
                return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        except:
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
    
    try:
        import pandas as pd
        import os
        from werkzeug.utils import secure_filename
        
        # 파일 업로드 확인
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': '파일이 업로드되지 않았습니다.'
            }), 400
        
        file = request.files['file']
        file_type = request.form.get('type', '')
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': '파일이 선택되지 않았습니다.'
            }), 400
        
        # 파일 확장자 확인
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'message': 'Excel 파일(.xlsx, .xls)만 업로드 가능합니다.'
            }), 400
        
        # 파일 저장
        filename = secure_filename(file.filename)
        upload_dir = os.path.join(os.getcwd(), 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        
        logger.info(f"파일 업로드 완료: {filename}, 타입: {file_type}")
        
        # 엑셀 파일 읽기
        try:
            df = pd.read_excel(file_path)
            logger.info(f"엑셀 파일 읽기 성공: {len(df)} 행")
        except Exception as e:
            logger.error(f"엑셀 파일 읽기 실패: {e}")
            return jsonify({
                'success': False,
                'message': f'엑셀 파일 읽기에 실패했습니다: {str(e)}'
            }), 500
        
        # 파일 타입에 따른 처리
        result = []
        
        if file_type == 'b-building':
            # B동전용 처리
            result = process_b_building_data(df)
        elif file_type == 'a-ground':
            # A동지상 (P2) 처리
            result = process_a_ground_data(df)
        elif file_type == 'a-basement':
            # A동지하 (P1) 처리
            result = process_a_basement_data(df)
        else:
            return jsonify({
                'success': False,
                'message': '잘못된 파일 타입입니다.'
            }), 400
        
        # 임시 파일 삭제
        try:
            os.remove(file_path)
        except:
            pass
        
        return jsonify({
            'success': True,
            'message': f'{len(result)}개의 항목이 정렬되었습니다.',
            'result': result
        })
        
    except Exception as e:
        logger.error(f"미출대응 파일 처리 에러: {e}")
        return jsonify({
            'success': False,
            'message': f'파일 처리 중 오류가 발생했습니다: {str(e)}'
        }), 500

def process_b_building_data(df):
    """B동전용 데이터 처리"""
    try:
        # 기본 정렬 (첫 번째 컬럼 기준)
        df_sorted = df.sort_values(df.columns[0])
        
        result = []
        for index, row in df_sorted.iterrows():
            result.append({
                'productCode': str(row.iloc[0]) if len(row) > 0 else '',
                'productName': str(row.iloc[1]) if len(row) > 1 else '',
                'quantity': str(row.iloc[2]) if len(row) > 2 else '',
                'location': 'B동'
            })
        
        return result
    except Exception as e:
        logger.error(f"B동 데이터 처리 에러: {e}")
        return []

def process_a_ground_data(df):
    """A동지상 (P2) 데이터 처리"""
    try:
        # P2 컬럼 찾기 (컬럼명에 'P2'가 포함된 컬럼)
        p2_column = None
        for col in df.columns:
            if 'P2' in str(col):
                p2_column = col
                break
        
        if p2_column is None:
            # P2 컬럼이 없으면 두 번째 컬럼 사용
            p2_column = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        
        # P2 기준으로 정렬
        df_sorted = df.sort_values(p2_column)
        
        result = []
        for index, row in df_sorted.iterrows():
            result.append({
                'productCode': str(row.iloc[0]) if len(row) > 0 else '',
                'productName': str(row.iloc[1]) if len(row) > 1 else '',
                'quantity': str(row.iloc[2]) if len(row) > 2 else '',
                'location': 'A동지상(P2)'
            })
        
        return result
    except Exception as e:
        logger.error(f"A동지상 데이터 처리 에러: {e}")
        return []

def process_a_basement_data(df):
    """A동지하 (P1) 데이터 처리"""
    try:
        # P1 컬럼 찾기 (컬럼명에 'P1'이 포함된 컬럼)
        p1_column = None
        for col in df.columns:
            if 'P1' in str(col):
                p1_column = col
                break
        
        if p1_column is None:
            # P1 컬럼이 없으면 첫 번째 컬럼 사용
            p1_column = df.columns[0]
        
        # P1 기준으로 정렬
        df_sorted = df.sort_values(p1_column)
        
        result = []
        for index, row in df_sorted.iterrows():
            result.append({
                'productCode': str(row.iloc[0]) if len(row) > 0 else '',
                'productName': str(row.iloc[1]) if len(row) > 1 else '',
                'quantity': str(row.iloc[2]) if len(row) > 2 else '',
                'location': 'A동지하(P1)'
            })
        
        return result
    except Exception as e:
        logger.error(f"A동지하 데이터 처리 에러: {e}")
        return []

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False) 